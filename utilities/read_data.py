import csv

from openpyxl import load_workbook
class Read_Data():
    def read_data_from_ex(file_name, sheet):
        datalist = []
        wb = load_workbook(filename = file_name)
        sh = wb[sheet]
        row_count = sh.max_row
        col_count = sh.max_column
        for i in range(2,row_count + 1):
            row = []
            for j in range(1,col_count + 1):
                row.append(sh.cell(row=i,column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        datalist = []
        csvdata = open(filename,"r")
        reader = csv.reader(csvdata)
        next(reader)
        for rows in reader:
            datalist.append(rows)
        return datalist
